import openpyxl as ox
import pandas as pd
from IPython.core.display import display
from random import randint
from random import random

student_id_list=[f"C0A21{i}"for i in range(101,201) ]
student_temp_list=[f"{randint(35,39)+round(random(),1)}"for _ in student_id_list]
time_list=["8:45:15" for _ in student_id_list]

acells_list=[f"A{number}"for number, _ in enumerate(student_id_list,1)]
bcells_list=[f"B{number}"for number, _ in enumerate(student_id_list,1)]
ccells_list=[f"C{number}"for number, _ in enumerate(student_id_list,1)]

student_id_list.insert(0,"学籍番号")
student_temp_list.insert(0,"体温")
time_list.insert(0,"受付時間")

wb=ox.Workbook()
ws=wb.active
for number,i in enumerate(student_id_list):
    temp_per_random=random()
    try:
        ws[acells_list[number]].value=i
        ws[bcells_list[number]].value=time_list[number]
        if temp_per_random<0.8:
            ws[ccells_list[number]].value=student_temp_list[number]
        else:
            ws[ccells_list[number]].value=student_temp_list[number]
    except:
        ws["A99"].value=i
        ws["B99"].value=time_list[number]
        if temp_per_random<0.9:
            ws["C99"].value=student_temp_list[number]
        else:
            ws["C99"].value=student_temp_list[number]

wb.save("demo_data.xlsx")

df=pd.read_excel("demo_data.xlsx")
"""
replace=df.iloc[97,2]
df.iloc[97,2]=df.iloc[97,1]
df.iloc[97,1]=replace
"""

df.to_csv("demo_data.csv",index=False)